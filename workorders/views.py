"""
Views for Work Orders app
"""
import io
import csv
from datetime import datetime
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, BasePermission
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import WorkOrder, WorkOrderComment, WorkOrderPart, WorkOrderStatus, WorkOrderType
from .serializers import (WorkOrderSerializer, WorkOrderListSerializer,
                          WorkOrderCommentSerializer, WorkOrderPartSerializer,
                          WorkOrderAssignSerializer)

User = get_user_model()


class IsAdminOrSupervisorOrReadOnly(BasePermission):
    """
    自定义权限：admin和supervisor可以修改，其他用户根据状态只读或操作
    """
    def has_permission(self, request, view):
        # 读取操作允许所有认证用户
        if request.method in ['GET', 'HEAD', 'OPTIONS']:
            return request.user.is_authenticated
        # 创建/更新/删除只允许admin或supervisor
        return request.user.is_authenticated and request.user.role in ['admin', 'supervisor']


class WorkOrderViewSet(viewsets.ModelViewSet):
    """ViewSet for WorkOrder model"""
    queryset = WorkOrder.objects.select_related(
        'equipment', 'requested_by', 'assignee', 'assigned_by',
        'maintenance_plan', 'completed_by', 'closed_by'
    ).prefetch_related('parts_used', 'comments')
    serializer_class = WorkOrderSerializer
    permission_classes = [IsAdminOrSupervisorOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['wo_type', 'status', 'priority', 'equipment', 'assignee']
    search_fields = ['wo_code', 'summary', 'description']
    ordering_fields = ['created_at', 'planned_start', 'priority']
    ordering = ['-created_at']

    def perform_create(self, serializer):
        """Set requested_by on create"""
        print(f"创建工单，用户: {self.request.user}, 数据: {serializer.validated_data}")
        serializer.save(requested_by=self.request.user)

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return WorkOrderListSerializer
        return WorkOrderSerializer

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def assign(self, request, pk=None):
        """Assign work order to a technician"""
        work_order = self.get_object()
        serializer = WorkOrderAssignSerializer(data=request.data)

        if serializer.is_valid():
            assignee_id = serializer.validated_data['assignee_id']

            # Check if work order can be assigned
            if not work_order.can_be_assigned():
                return Response(
                    {"error": f"Work order cannot be assigned in current status: {work_order.status}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                assignee = User.objects.get(id=assignee_id)
            except User.DoesNotExist:
                return Response(
                    {"error": "Assignee not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            work_order.assignee = assignee
            work_order.assigned_by = request.user
            work_order.assigned_at = timezone.now()
            work_order.status = WorkOrderStatus.ASSIGNED
            work_order.save()

            # Log audit
            from users.models import AuditLog
            AuditLog.log(
                actor=request.user,
                action='assign',
                entity_type='WorkOrder',
                entity_id=work_order.id,
                entity_repr=str(work_order)
            )

            return Response(WorkOrderSerializer(work_order).data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def start(self, request, pk=None):
        """Start work on a work order"""
        work_order = self.get_object()

        if not work_order.can_be_started():
            return Response(
                {"error": f"Work order cannot be started in current status: {work_order.status}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        work_order.status = WorkOrderStatus.IN_PROGRESS
        work_order.actual_start = timezone.now()
        work_order.save()

        # Log audit
        from users.models import AuditLog
        AuditLog.log(
            actor=request.user,
            action='update',
            entity_type='WorkOrder',
            entity_id=work_order.id,
            entity_repr=str(work_order),
            diff={'status': WorkOrderStatus.IN_PROGRESS}
        )

        return Response(WorkOrderSerializer(work_order).data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def complete(self, request, pk=None):
        """Complete a work order"""
        work_order = self.get_object()

        if not work_order.can_be_completed():
            return Response(
                {"error": f"Work order cannot be completed in current status: {work_order.status}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update work order with completion data if provided
        actions_taken = request.data.get('actions_taken')
        if actions_taken:
            work_order.actions_taken = actions_taken
        
        # Update other completion fields if provided
        if 'root_cause' in request.data:
            work_order.root_cause = request.data['root_cause']
        if 'downtime_minutes' in request.data:
            work_order.downtime_minutes = request.data['downtime_minutes']
        if 'labor_hours' in request.data:
            work_order.labor_hours = request.data['labor_hours']
        if 'parts_cost' in request.data:
            work_order.parts_cost = request.data['parts_cost']
        if 'notes' in request.data:
            work_order.notes = request.data['notes']

        # Validate required fields
        if not work_order.actions_taken:
            return Response(
                {"error": "actions_taken is required to complete a work order"},
                status=status.HTTP_400_BAD_REQUEST
            )

        work_order.status = WorkOrderStatus.COMPLETED
        work_order.actual_end = timezone.now()
        work_order.completed_by = request.user
        work_order.completed_at = timezone.now()
        work_order.save()

        # Log audit
        from users.models import AuditLog
        AuditLog.log(
            actor=request.user,
            action='complete',
            entity_type='WorkOrder',
            entity_id=work_order.id,
            entity_repr=str(work_order)
        )

        return Response(WorkOrderSerializer(work_order).data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def close(self, request, pk=None):
        """Close a work order"""
        work_order = self.get_object()

        if not work_order.can_be_closed():
            return Response(
                {"error": f"Work order cannot be closed in current status: {work_order.status}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        work_order.status = WorkOrderStatus.CLOSED
        work_order.closed_by = request.user
        work_order.closed_at = timezone.now()
        work_order.save()

        # Log audit
        from users.models import AuditLog
        AuditLog.log(
            actor=request.user,
            action='close',
            entity_type='WorkOrder',
            entity_id=work_order.id,
            entity_repr=str(work_order)
        )

        return Response(WorkOrderSerializer(work_order).data)

    @action(detail=False, methods=['get'])
    def my_orders(self, request):
        """Get work orders assigned to current user"""
        my_orders = self.get_queryset().filter(assignee=request.user)
        serializer = WorkOrderListSerializer(my_orders, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue work orders"""
        overdue_orders = [wo for wo in self.get_queryset() if wo.is_overdue]
        serializer = WorkOrderListSerializer(overdue_orders, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download Excel template for work order import"""
        wb = Workbook()
        ws = wb.active
        ws.title = "工单导入模板"

        # 表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # 定义表头
        headers = [
            ('equipment_code', '设备编码*', '必填'),
            ('summary', '工单摘要*', '必填'),
            ('description', '工单描述', ''),
            ('wo_type', '工单类型', 'PM/CM/inspection'),
            ('priority', '优先级', 'low/medium/high/critical'),
            ('planned_start', '计划开始时间', '格式: YYYY-MM-DD HH:MM'),
            ('planned_end', '计划结束时间', '格式: YYYY-MM-DD HH:MM'),
            ('failure_code', '故障代码', ''),
            ('notes', '备注', ''),
        ]

        # 写入表头
        for col_idx, (field, name, desc) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = f"{name}\n({desc})"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # 设置列宽
        column_widths = [18, 25, 30, 15, 15, 20, 20, 15, 30]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # 添加示例数据
        example_data = [
            'EQ-001', '设备异响维修', '运行时有异响，需要检查', 'CM', 'high',
            '2024-01-15 09:00', '2024-01-15 17:00', 'M-001', '紧急维修'
        ]
        for col_idx, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col_idx, value=value)

        # 设置行高
        ws.row_dimensions[1].height = 40

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="工单导入模板.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import work orders from Excel or CSV file"""
        from assets.models import Asset

        file = request.FILES.get('file')
        if not file:
            return Response(
                {"error": "请选择要导入的文件"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 检查文件类型
        file_ext = file.name.split('.')[-1].lower()
        if file_ext not in ['xlsx', 'xls', 'csv']:
            return Response(
                {"error": "仅支持 Excel (.xlsx, .xls) 或 CSV 文件"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # 读取文件
            if file_ext == 'csv':
                data = self._read_csv_file(file)
            else:
                data = self._read_excel_file(file)

            # 处理导入
            result = self._process_import_data(data, request.user)

            return Response(result, status=status.HTTP_200_OK if result['success_count'] > 0 else status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response(
                {"error": f"导入失败: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _read_csv_file(self, file):
        """Read CSV file and return data list"""
        data = []
        decoded_file = file.read().decode('utf-8-sig')
        reader = csv.DictReader(decoded_file.splitlines())

        # 获取中文列名到字段的映射
        header_map = self._get_header_map(reader.fieldnames)

        for row_num, row in enumerate(reader, start=2):
            data.append({
                'row_num': row_num,
                'data': self._map_row_data(row, header_map)
            })
        return data

    def _read_excel_file(self, file):
        """Read Excel file and return data list"""
        import openpyxl
        data = []
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # 获取表头映射
        headers = []
        for cell in ws[1]:
            value = cell.value
            if value:
                # 从"中文名 (说明)"格式中提取中文名
                header_name = str(value).split('\n')[0].split('(')[0].strip('*').strip()
                headers.append(header_name)
            else:
                headers.append('')

        header_map = self._get_header_map(headers)

        # 读取数据行（跳过第一行表头）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if any(cell is not None for cell in row):
                row_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        field_name = header_map.get(headers[col_idx])
                        if field_name:
                            row_data[field_name] = value
                data.append({
                    'row_num': row_idx,
                    'data': row_data
                })

        return data

    def _get_header_map(self, headers):
        """Get mapping from Chinese header names to field names"""
        header_map = {
            '设备编码': 'equipment_code',
            '工单摘要': 'summary',
            '工单描述': 'description',
            '工单类型': 'wo_type',
            '优先级': 'priority',
            '计划开始时间': 'planned_start',
            '计划结束时间': 'planned_end',
            '故障代码': 'failure_code',
            '备注': 'notes',
        }
        # 同时支持英文字段名
        for field in header_map.values():
            header_map[field] = field
        return header_map

    def _map_row_data(self, row, header_map):
        """Map CSV row data to field names"""
        mapped_data = {}
        for csv_header, value in row.items():
            if csv_header in header_map:
                field_name = header_map[csv_header]
                mapped_data[field_name] = value
        return mapped_data

    def _process_import_data(self, data, user):
        """Process imported data and create work orders"""
        from assets.models import Asset

        success_count = 0
        error_count = 0
        errors = []

        for item in data:
            row_num = item['row_num']
            wo_data = item['data']

            try:
                # 验证必填字段
                equipment_code = wo_data.get('equipment_code')
                if not equipment_code:
                    errors.append(f"第{row_num}行: 设备编码不能为空")
                    error_count += 1
                    continue

                summary = wo_data.get('summary')
                if not summary:
                    errors.append(f"第{row_num}行: 工单摘要不能为空")
                    error_count += 1
                    continue

                # 查找设备
                try:
                    equipment = Asset.objects.get(code=equipment_code)
                except Asset.DoesNotExist:
                    errors.append(f"第{row_num}行: 设备编码 '{equipment_code}' 不存在")
                    error_count += 1
                    continue

                # 处理日期时间字段
                for date_field in ['planned_start', 'planned_end']:
                    value = wo_data.get(date_field)
                    if value:
                        if isinstance(value, datetime):
                            wo_data[date_field] = value
                        elif isinstance(value, str):
                            try:
                                wo_data[date_field] = datetime.strptime(value, '%Y-%m-%d %H:%M')
                            except ValueError:
                                try:
                                    wo_data[date_field] = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                                except ValueError:
                                    try:
                                        wo_data[date_field] = datetime.strptime(value, '%Y-%m-%d')
                                    except ValueError:
                                        errors.append(f"第{row_num}行: {date_field} 日期时间格式错误，应为 YYYY-MM-DD HH:MM")
                                        error_count += 1
                                        continue

                # 设置默认值
                wo_data.setdefault('wo_type', 'CM')
                wo_data.setdefault('priority', 'medium')
                wo_data['equipment'] = equipment
                wo_data['requested_by'] = user

                # 清理不需要的字段
                wo_data.pop('equipment_code', None)

                # 创建工单
                WorkOrder.objects.create(**wo_data)
                success_count += 1

            except Exception as e:
                errors.append(f"第{row_num}行: {str(e)}")
                error_count += 1

        return {
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:20],  # 最多返回20条错误
            'message': f'导入完成: 成功 {success_count} 条, 失败 {error_count} 条'
        }

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export work orders to Excel file"""
        work_orders = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "工单列表"

        # 表头
        headers = [
            '工单号', '设备编码', '设备名称', '工单类型', '摘要', '描述',
            '优先级', '状态', '负责人', '计划开始', '计划结束',
            '实际开始', '实际结束', '故障代码', '根本原因',
            '停机时间(分钟)', '人工工时', '备件成本', '备注', '创建时间'
        ]

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for row_idx, wo in enumerate(work_orders, 2):
            ws.cell(row=row_idx, column=1, value=wo.wo_code or '')
            ws.cell(row=row_idx, column=2, value=wo.equipment.code if wo.equipment else '')
            ws.cell(row=row_idx, column=3, value=wo.equipment.name if wo.equipment else '')
            ws.cell(row=row_idx, column=4, value=wo.get_wo_type_display() or '')
            ws.cell(row=row_idx, column=5, value=wo.summary or '')
            ws.cell(row=row_idx, column=6, value=wo.description or '')
            ws.cell(row=row_idx, column=7, value=wo.get_priority_display() or '')
            ws.cell(row=row_idx, column=8, value=wo.get_status_display() or '')
            ws.cell(row=row_idx, column=9, value=wo.assignee.full_name if wo.assignee else '')
            ws.cell(row=row_idx, column=10, value=wo.planned_start.strftime('%Y-%m-%d %H:%M') if wo.planned_start else '')
            ws.cell(row=row_idx, column=11, value=wo.planned_end.strftime('%Y-%m-%d %H:%M') if wo.planned_end else '')
            ws.cell(row=row_idx, column=12, value=wo.actual_start.strftime('%Y-%m-%d %H:%M') if wo.actual_start else '')
            ws.cell(row=row_idx, column=13, value=wo.actual_end.strftime('%Y-%m-%d %H:%M') if wo.actual_end else '')
            ws.cell(row=row_idx, column=14, value=wo.failure_code or '')
            ws.cell(row=row_idx, column=15, value=wo.root_cause or '')
            ws.cell(row=row_idx, column=16, value=wo.downtime_minutes or 0)
            ws.cell(row=row_idx, column=17, value=float(wo.labor_hours) if wo.labor_hours else 0)
            ws.cell(row=row_idx, column=18, value=float(wo.parts_cost) if wo.parts_cost else 0)
            ws.cell(row=row_idx, column=19, value=wo.notes or '')
            ws.cell(row=row_idx, column=20, value=wo.created_at.strftime('%Y-%m-%d %H:%M') if wo.created_at else '')

        # 设置列宽
        for col_idx in range(1, 21):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 15

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="工单列表_{timestamp}.xlsx"'
        return response


class WorkOrderCommentViewSet(viewsets.ModelViewSet):
    """ViewSet for WorkOrderComment model"""
    queryset = WorkOrderComment.objects.select_related('author', 'work_order').all()
    serializer_class = WorkOrderCommentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['work_order', 'is_internal']

    def perform_create(self, serializer):
        """Set author on create"""
        serializer.save(author=self.request.user)


class WorkOrderPartViewSet(viewsets.ModelViewSet):
    """ViewSet for WorkOrderPart model"""
    queryset = WorkOrderPart.objects.select_related('work_order').all()
    serializer_class = WorkOrderPartSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['work_order']
