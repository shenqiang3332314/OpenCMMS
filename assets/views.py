"""
Views for Assets app
"""
import io
import csv
import openpyxl
from datetime import datetime
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, BasePermission
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Asset, AssetStatus
from .serializers import AssetSerializer, AssetListSerializer, AssetTreeSerializer


class AssetPagination(PageNumberPagination):
    """自定义分页类，支持更大的页面大小"""
    page_size = 50  # 默认每页50条
    page_size_query_param = 'page_size'  # 允许客户端通过page_size参数控制
    max_page_size = 1000  # 最大允许1000条，足够显示所有设备


class IsAdminOrReadOnly(BasePermission):
    """
    自定义权限：只有admin可以修改，其他用户只读
    """
    def has_permission(self, request, view):
        # 读取操作允许所有认证用户
        if request.method in ['GET', 'HEAD', 'OPTIONS']:
            return request.user.is_authenticated
        # 写入操作只允许admin
        return request.user.is_authenticated and request.user.role == 'admin'


class AssetViewSet(viewsets.ModelViewSet):
    """ViewSet for Asset model"""
    queryset = Asset.objects.select_related('parent', 'created_by').all()
    serializer_class = AssetSerializer
    permission_classes = [IsAdminOrReadOnly]
    pagination_class = AssetPagination  # 使用自定义分页类
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'factory', 'workshop', 'line', 'criticality']
    search_fields = ['code', 'name', 'equipment_id', 'serial_number', 'vendor']
    ordering_fields = ['code', 'name', 'created_at']
    ordering = ['code']

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return AssetListSerializer
        return AssetSerializer

    def perform_create(self, serializer):
        """Set created_by on create"""
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get assets overdue for maintenance"""
        overdue_assets = [a for a in self.get_queryset() if a.is_overdue_for_maintenance()]
        serializer = AssetListSerializer(overdue_assets, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def tree(self, request):
        """Get asset hierarchy tree"""
        root_assets = self.get_queryset().filter(parent__isnull=True)
        serializer = AssetTreeSerializer(root_assets, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def children(self, request, pk=None):
        """Get child assets"""
        asset = self.get_object()
        children = asset.children.all()
        serializer = AssetListSerializer(children, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download Excel template for asset import"""
        wb = Workbook()
        ws = wb.active
        ws.title = "设备导入模板"

        # 表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # 定义表头
        headers = [
            ('code', '设备编码*', '必填，唯一'),
            ('name', '设备名称*', '必填'),
            ('process', '工艺', ''),
            ('equipment_id', '设备ID', ''),
            ('machine_name', '机器名称', ''),
            ('factory', '工厂', ''),
            ('workshop', '车间', ''),
            ('line', '产线', ''),
            ('station', '工位', ''),
            ('vendor', '供应商', ''),
            ('model', '型号', ''),
            ('serial_number', '序列号', ''),
            ('specification', '规格说明', ''),
            ('start_date', '投用日期', '格式: YYYY-MM-DD'),
            ('warranty_expiry', '保修到期', '格式: YYYY-MM-DD'),
            ('status', '状态', 'active/inactive/retired/maintenance'),
            ('criticality', '重要性', 'critical/important/normal'),
            ('cost_center', '成本中心', ''),
            ('asset_value', '资产价值', '数字'),
            ('expected_life_years', '预计使用年限(年)', '数字'),
            ('meter_unit', '计量单位', ''),
            ('current_meter_reading', '当前计量读数', '数字'),
            ('notes', '备注', ''),
        ]

        # 写入表头
        for col_idx, (field, name, desc) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = f"{name}\n({desc})"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # 设置列宽
        column_widths = [15, 20, 12, 15, 15, 12, 12, 12, 12, 15, 15, 18, 30, 15, 15, 20, 18, 12, 12, 15, 12, 18, 30]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # 添加示例数据
        example_data = [
            'EQ-001', '数控机床', '机加工', 'EID-001', 'CNC-01', '一厂', '机加车间',
            'Line-1', 'Station-A', 'Siemens', 'X2000', 'SN202401001', '额定功率10kW',
            '2024-01-15', '2025-01-15', 'active', 'important', 'CC001', '150000.00', '10', 'hours', '1000', '设备运行正常'
        ]
        for col_idx, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col_idx, value=value)

        # 设置行高
        ws.row_dimensions[1].height = 40

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="设备导入模板.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import assets from Excel or CSV file"""
        file = request.FILES.get('file')
        if not file:
            return Response(
                {"error": "请选择要导入的文件"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 检查文件类型
        file_ext = file.name.split('.')[-1].lower()
        if file_ext not in ['xlsx', 'xls', 'csv']:
            return Response(
                {"error": "仅支持 Excel (.xlsx, .xls) 或 CSV 文件"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # 读取文件
            if file_ext == 'csv':
                data = self._read_csv_file(file)
            else:
                data = self._read_excel_file(file)

            # 处理导入
            result = self._process_import_data(data, request.user)

            return Response(result, status=status.HTTP_200_OK if result['success_count'] > 0 else status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response(
                {"error": f"导入失败: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _read_csv_file(self, file):
        """Read CSV file and return data list"""
        data = []
        decoded_file = file.read().decode('utf-8-sig')
        reader = csv.DictReader(decoded_file.splitlines())

        # 获取中文列名到字段的映射
        header_map = self._get_header_map(reader.fieldnames)

        for row_num, row in enumerate(reader, start=2):
            data.append({
                'row_num': row_num,
                'data': self._map_row_data(row, header_map)
            })
        return data

    def _read_excel_file(self, file):
        """Read Excel file and return data list"""
        data = []
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # 获取表头映射
        headers = []
        for cell in ws[1]:
            value = cell.value
            if value:
                # 从"中文名 (说明)"格式中提取中文名
                header_name = str(value).split('\n')[0].split('(')[0].strip('*').strip()
                headers.append(header_name)
            else:
                headers.append('')

        header_map = self._get_header_map(headers)

        # 读取数据行（跳过第一行表头）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if any(cell is not None for cell in row):
                row_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        field_name = header_map.get(headers[col_idx])
                        if field_name:
                            row_data[field_name] = value
                data.append({
                    'row_num': row_idx,
                    'data': row_data
                })

        return data

    def _get_header_map(self, headers):
        """Get mapping from Chinese header names to field names"""
        header_map = {
            '设备编码': 'code',
            '设备名称': 'name',
            '工艺': 'process',
            '设备ID': 'equipment_id',
            '机器名称': 'machine_name',
            '工厂': 'factory',
            '车间': 'workshop',
            '产线': 'line',
            '工位': 'station',
            '供应商': 'vendor',
            '型号': 'model',
            '序列号': 'serial_number',
            '规格说明': 'specification',
            '投用日期': 'start_date',
            '保修到期': 'warranty_expiry',
            '状态': 'status',
            '重要性': 'criticality',
            '成本中心': 'cost_center',
            '资产价值': 'asset_value',
            '预计使用年限(年)': 'expected_life_years',
            '计量单位': 'meter_unit',
            '当前计量读数': 'current_meter_reading',
            '备注': 'notes',
        }
        # 同时支持英文字段名 - 修复字典遍历时修改的问题
        english_fields = {}
        for field in header_map.values():
            english_fields[field] = field
        header_map.update(english_fields)
        return header_map

    def _map_row_data(self, row, header_map):
        """Map CSV row data to field names"""
        mapped_data = {}
        for csv_header, value in row.items():
            if csv_header in header_map:
                field_name = header_map[csv_header]
                mapped_data[field_name] = value
        return mapped_data

    def _process_import_data(self, data, user):
        """Process imported data and create assets"""
        success_count = 0
        error_count = 0
        errors = []

        for item in data:
            row_num = item['row_num']
            asset_data = item['data']

            try:
                # 验证必填字段
                if not asset_data.get('code'):
                    errors.append(f"第{row_num}行: 设备编码不能为空")
                    error_count += 1
                    continue

                if not asset_data.get('name'):
                    errors.append(f"第{row_num}行: 设备名称不能为空")
                    error_count += 1
                    continue

                # 检查编码是否已存在
                if Asset.objects.filter(code=asset_data['code']).exists():
                    errors.append(f"第{row_num}行: 设备编码 '{asset_data['code']}' 已存在")
                    error_count += 1
                    continue

                # 处理日期字段
                for date_field in ['start_date', 'warranty_expiry']:
                    value = asset_data.get(date_field)
                    if value:
                        if isinstance(value, datetime):
                            asset_data[date_field] = value.date()
                        elif isinstance(value, str):
                            try:
                                asset_data[date_field] = datetime.strptime(value, '%Y-%m-%d').date()
                            except ValueError:
                                try:
                                    asset_data[date_field] = datetime.strptime(value, '%Y/%m/%d').date()
                                except ValueError:
                                    errors.append(f"第{row_num}行: {date_field} 日期格式错误，应为 YYYY-MM-DD")
                                    error_count += 1
                                    continue

                # 设置默认值
                asset_data.setdefault('status', 'active')
                asset_data.setdefault('criticality', 'normal')
                asset_data.setdefault('current_meter_reading', 0)
                asset_data['created_by'] = user

                # 创建设备
                Asset.objects.create(**asset_data)
                success_count += 1

            except Exception as e:
                errors.append(f"第{row_num}行: {str(e)}")
                error_count += 1

        return {
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:20],  # 最多返回20条错误
            'message': f'导入完成: 成功 {success_count} 条, 失败 {error_count} 条'
        }

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export assets to Excel file"""
        assets = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "设备台账"

        # 表头
        headers = [
            '设备编码', '设备名称', '工艺', '设备ID', '机器名称', '工厂', '车间',
            '产线', '工位', '供应商', '型号', '序列号', '规格说明', '投用日期',
            '保修到期', '状态', '重要性', '成本中心', '资产价值', '预计使用年限',
            '计量单位', '当前计量读数', '最后保养日期', '下次保养日期', '备注'
        ]

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for row_idx, asset in enumerate(assets, 2):
            ws.cell(row=row_idx, column=1, value=asset.code)
            ws.cell(row=row_idx, column=2, value=asset.name)
            ws.cell(row=row_idx, column=3, value=asset.process or '')
            ws.cell(row=row_idx, column=4, value=asset.equipment_id or '')
            ws.cell(row=row_idx, column=5, value=asset.machine_name or '')
            ws.cell(row=row_idx, column=6, value=asset.factory or '')
            ws.cell(row=row_idx, column=7, value=asset.workshop or '')
            ws.cell(row=row_idx, column=8, value=asset.line or '')
            ws.cell(row=row_idx, column=9, value=asset.station or '')
            ws.cell(row=row_idx, column=10, value=asset.vendor or '')
            ws.cell(row=row_idx, column=11, value=asset.model or '')
            ws.cell(row=row_idx, column=12, value=asset.serial_number or '')
            ws.cell(row=row_idx, column=13, value=asset.specification or '')
            ws.cell(row=row_idx, column=14, value=asset.start_date.strftime('%Y-%m-%d') if asset.start_date else '')
            ws.cell(row=row_idx, column=15, value=asset.warranty_expiry.strftime('%Y-%m-%d') if asset.warranty_expiry else '')
            ws.cell(row=row_idx, column=16, value=asset.status)
            ws.cell(row=row_idx, column=17, value=asset.criticality)
            ws.cell(row=row_idx, column=18, value=asset.cost_center or '')
            ws.cell(row=row_idx, column=19, value=float(asset.asset_value) if asset.asset_value else '')
            ws.cell(row=row_idx, column=20, value=asset.expected_life_years or '')
            ws.cell(row=row_idx, column=21, value=asset.meter_unit or '')
            ws.cell(row=row_idx, column=22, value=float(asset.current_meter_reading))
            ws.cell(row=row_idx, column=23, value=asset.last_maintenance_date.strftime('%Y-%m-%d') if asset.last_maintenance_date else '')
            ws.cell(row=row_idx, column=24, value=asset.next_maintenance_date.strftime('%Y-%m-%d') if asset.next_maintenance_date else '')
            ws.cell(row=row_idx, column=25, value=asset.notes or '')

        # 设置列宽
        for col_idx in range(1, 26):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 15

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="设备台账_{timestamp}.xlsx"'
        return response
